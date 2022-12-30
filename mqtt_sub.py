import xlsxwriter
import ssl
import paho.mqtt.client as paho
import paho.mqtt.subscribe as subscribe
from paho import mqtt
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
import pandas as pd
import os.path
# pip install openpyxl

# pip install pandas

brokerUrl = "821032442a5c4b30982ba8c1b2085d07.s2.eu.hivemq.cloud"
brokerPort = 8883
brokerUsername = "morgana"
brokerPassword = "morgana01"
_filename = 'dash10.xlsx'


if os.path.exists(_filename):
  os.remove(_filename)

def gerarGrafico(filename):
    workbook = xlsxwriter.Workbook(filename) 
    worksheet = workbook.add_worksheet('Grafico')

    wb = load_workbook(_filename)
    sh = wb.active
    row = str(sh.max_row)
    print(row)

    chart1 = workbook.add_chart({'type': 'line'}) 
    chart1.set_up_down_bars() 
    
    chart1.add_series({  
        'categories': '= Sheet1 !$A$2:$A${row}',  
        'values':     '= Sheet1 !$B$2:$B${row}',  
    })  
        
    chart1.add_series({  
        'categories': '= Sheet1 !$A$2:$A${row}',  
        'values':     '= Sheet1 !$C$2:$C${row}',  
    })  
        
    chart1.set_title({'name': 'Dash IOT'})  
    chart1.set_x_axis({'name': 'Subjects'})   
    chart1.set_y_axis({'name': 'Marks'})  
    chart1.set_style(11)  
    worksheet.insert_chart('D2', chart1,   
        {'x_offset': 20, 'y_offset': 5})  
    workbook.close() 

def caso_arquivo_nao_exista_cria_ele(filename):
    if os.path.exists(filename) == 0:
        wb=Workbook()
        wb.save(filename)
        wb = load_workbook(filename)
        sh = wb.active
        sh['A1'] = 'Temperatura'
        sh['B1'] = 'Presença'
        font = Font(bold=True)
        sh['A1'].font = font
        sh['B1'].font = font
        wb.save(filename)

# callback to print a message once it arrives 
def print_msg(client, userdata, message):
    """
        Prints a mqtt message to stdout ( used as callback for subscribe )
        :param client: the client itself
        :param userdata: userdata is set when initiating the client, here it is userdata=None
        :param message: the message with topic and payload
    """
    print("%s : %s" % (message.topic, message.payload))
    if message.topic == 'lab/temperature':
        caso_arquivo_nao_exista_cria_ele(_filename)
        wb = load_workbook(_filename)
        sh = wb.active
        row = str(sh.max_row+1)
        # print(sh['A16'].value)
        val1='A'+row
        # val2='B'+row
        # load demo.xlsx 
        wb=load_workbook(_filename)
        # select demo.xlsx
        sheet=wb.active
        # set value for cell A1=1
        sheet[val1] = message.payload
        # sheet[val2] = message.payload
        # set value for cell B2=2
        # sheet.cell(row=2, column=2).value = 2
        # save workbook 
        wb.save(_filename)
    if message.topic == 'lab/presence':
        wb = load_workbook(_filename)
        sh = wb.active
        row = str(sh.max_row)
        val1='B'+row
        # val2='D'+row
        # load demo.xlsx 
        wb=load_workbook(_filename)
        # select demo.xlsx
        sheet=wb.active
        # set value for cell A1=1
        sheet[val1] = message.payload
        # sheet[val2] = message.payload
        # set value for cell B2=2
        # sheet.cell(row=2, column=2).value = 2
        # save workbook 
        wb.save(_filename)
        # gerarGrafico(_filename)


# sheet = load_workbook(filename = '/home/aline/Área de Trabalho/Pecris/IFPE/IoT/hello.xlsx')
# ow_count = sheet.max_row
# print(ow_count)
# df = pd.read_excel(r"/home/aline/Área de Trabalho/Pecris/IFPE/IoT/asd.xlsx")
# # qtd_linhas = df.count()
# print(df)


# use TLS for secure connection with HiveMQ Cloud
sslSettings = ssl.SSLContext(mqtt.client.ssl.PROTOCOL_TLS)

# put in your cluster credentials and hostname
auth = {'username': brokerUsername, 'password': brokerPassword}
subscribe.callback(print_msg, "#", hostname=brokerUrl, port=brokerPort, auth=auth,
                   tls=sslSettings, protocol=paho.MQTTv31)